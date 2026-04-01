from fastapi import FastAPI, Request, Query, Form
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
import psycopg2
from psycopg2.extras import RealDictCursor
import pandas as pd
import os
import uvicorn
import secrets
from starlette.middleware.sessions import SessionMiddleware
from fastapi.responses import FileResponse
import smtplib
from email.message import EmailMessage
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STATIC_DIR = os.path.join(BASE_DIR, "static")
TEMPLATES_DIR = os.path.join(BASE_DIR, "templates")

app = FastAPI()

SECRET_KEY = os.getenv("SESSION_SECRET", secrets.token_hex(32))

app.add_middleware(
    SessionMiddleware,
    secret_key=SECRET_KEY,
    max_age=60*60*8,
    same_site="lax"
)

templates = Jinja2Templates(directory=TEMPLATES_DIR)
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

VAT_RATE = 0.15

# ---------------- USERS ----------------

USERS = {
    "ryan":{"password":"@Karabo2009@","role":"owner"},
    "lebo":{"password":"Karabo@2009","role":"accounts"},
    "admin":{"password":"Malebo2913$","role":"admin"}
}

# ---------------- PARTIES ----------------

PARTIES = [
"KONE","OTIS","ALICEWEAR","SPIRAX SARCO","TRACLO PTY LTD",
"TRACLO INTL","TRACLO INTER","MAXIONWHEEL","MINTEK",
"YMS TRADING DISTRIBUTORS","WALK-IN","WEG","SULZER",
"CUSTOMS","USAFETY","SILVER","MAXION",
"Mahniglory and Saama PTY LTD",
"UPPER LEVEL LIFTS PTY LTD",
"Power Elevators",
"Imperio Logistics Holdings (Pty) Ltd"
]

SUPPLIERS = ["DHL","JKJ","MOK"]

# ---------------- DATABASE ----------------

def get_conn():
    return psycopg2.connect(
        os.getenv("DATABASE_URL"),
        cursor_factory=RealDictCursor
    )

def init_db():

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS sales(
        id SERIAL PRIMARY KEY,
        party TEXT,
        supplier TEXT,
        order_no TEXT,
        invoice_no TEXT,
        waybill TEXT,
        sale_date DATE,
        supplier_cost NUMERIC,
        client_charge NUMERIC,
        vat NUMERIC,
        total_invoice NUMERIC,
        profit NUMERIC,
        paid_status TEXT
    )
    """)

    conn.commit()
    cur.close()
    conn.close()

init_db()

# ---------------- MODEL ----------------

class Sale(BaseModel):

    party:str
    supplier:str
    order_no:str
    invoice_no:str
    waybill:str
    sale_date:str
    supplier_cost:float
    client_charge:float
    paid_status:str

# ---------------- ROLE HELPER ----------------

def require_role(request:Request,roles):

    role=request.session.get("role")

    if role not in roles:
        return False

    return True

# ---------------- LOGIN ----------------

@app.get("/login",response_class=HTMLResponse)
def login_page(request:Request):

    return templates.TemplateResponse(
    request=request,
    name="login.html",
    context={}
)

@app.post("/login")
def login(request:Request,username:str=Form(...),password:str=Form(...)):

    if username in USERS and USERS[username]["password"]==password:

        request.session["user"]=username
        request.session["role"]=USERS[username]["role"]

        if USERS[username]["role"]=="owner":
            return RedirectResponse("/owner-dashboard",302)

        return RedirectResponse("/",302)

    return templates.TemplateResponse(
    request=request,
    name="login.html",
    context={"error": "Invalid login"}
)

@app.get("/logout")
def logout(request:Request):

    request.session.clear()
    return RedirectResponse("/login")

# ---------------- HOME ----------------

@app.get("/",response_class=HTMLResponse)
def home(request:Request):

    if not request.session.get("user"):
        return RedirectResponse("/login")


    return templates.TemplateResponse(
       request=request,
       name="index.html",
       context={
          "parties": PARTIES,
          "suppliers": SUPPLIERS,
          "role": request.session.get("role")
               }
    )   


# ---------------- RECORD SALE ----------------

@app.post("/record-sale")
def record_sale(request:Request,sale:Sale,vat_enabled:bool=Query(True)):

    if not require_role(request,["accounts","admin"]):
        return {"error":"Permission denied"}

    vat=sale.client_charge*VAT_RATE if vat_enabled else 0
    total=sale.client_charge+vat
    profit=sale.client_charge-sale.supplier_cost

    sale_date=sale.sale_date if sale.sale_date else None

    conn=get_conn()
    cur=conn.cursor()

    cur.execute("""
    INSERT INTO sales(
    party,supplier,order_no,invoice_no,waybill,sale_date,
    supplier_cost,client_charge,vat,total_invoice,profit,paid_status
    )
    VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """,(
        sale.party,
        sale.supplier,
        sale.order_no,
        sale.invoice_no,
        sale.waybill,
        sale_date,
        sale.supplier_cost,
        sale.client_charge,
        vat,
        total,
        profit,
        sale.paid_status
    ))

    conn.commit()
    cur.close()
    conn.close()

    return {
        "status":"recorded",
        "vat":vat,
        "total_invoice":total,
        "profit":profit
    }

# ---------------- UPDATE SALE ----------------

@app.put("/update-sale/{sale_id}")
def update_sale(request:Request,sale_id:int,sale:Sale,vat_enabled:bool=Query(True)):

    if not require_role(request,["accounts","admin"]):
        return {"error":"Permission denied"}

    vat=sale.client_charge*VAT_RATE if vat_enabled else 0
    total=sale.client_charge+vat
    profit=sale.client_charge-sale.supplier_cost

    sale_date=sale.sale_date if sale.sale_date else None

    conn=get_conn()
    cur=conn.cursor()

    cur.execute("""
    UPDATE sales SET
    party=%s,
    supplier=%s,
    order_no=%s,
    invoice_no=%s,
    waybill=%s,
    sale_date=%s,
    supplier_cost=%s,
    client_charge=%s,
    vat=%s,
    total_invoice=%s,
    profit=%s,
    paid_status=%s
    WHERE id=%s
    """,(
        sale.party,
        sale.supplier,
        sale.order_no,
        sale.invoice_no,
        sale.waybill,
        sale_date,
        sale.supplier_cost,
        sale.client_charge,
        vat,
        total,
        profit,
        sale.paid_status,
        sale_id
    ))

    conn.commit()
    cur.close()
    conn.close()

    return {"status":"updated"}

# ---------------- DELETE SALE ----------------

@app.delete("/delete-sale/{sale_id}")
def delete_sale(request:Request,sale_id:int):

    if not require_role(request,["accounts","admin"]):
        return {"error":"Permission denied"}

    conn=get_conn()
    cur=conn.cursor()

    cur.execute("DELETE FROM sales WHERE id=%s",(sale_id,))

    conn.commit()
    cur.close()
    conn.close()

    return {"status":"deleted"}

# ---------------- SALES TABLE ----------------

@app.get("/sales")
def get_sales():

    conn=get_conn()
    cur=conn.cursor()

    cur.execute("SELECT * FROM sales ORDER BY id DESC")

    rows=cur.fetchall()

    cur.close()
    conn.close()

    return rows


# ----------------Search SALES  ----------------
@app.get("/search-sales")
def search_sales(q: str = ""):

    conn = get_conn()
    cur = conn.cursor()

    query = """
    SELECT * FROM sales
    WHERE
        LOWER(party) LIKE %s OR
        LOWER(invoice_no) LIKE %s OR
        LOWER(order_no) LIKE %s
    ORDER BY id DESC
    """

    search = f"%{q.lower()}%"

    cur.execute(query, (search, search, search))
    rows = cur.fetchall()

    cur.close()
    conn.close()

    return rows

# ---------------- Client Statement ----------------
@app.get("/client-statement")
def client_statement(party: str, month: str, view: str = "internal"):

    conn = get_conn()

    query = """
        SELECT invoice_no, sale_date, client_charge, profit, paid_status
        FROM sales
        WHERE LOWER(party) = LOWER(%s)
        AND sale_date IS NOT NULL
        AND TO_CHAR(sale_date,'YYYY-MM') = %s
        ORDER BY sale_date
    """

    df = pd.read_sql(query, conn, params=(party, month))
    conn.close()

    if df.empty:
        return {"error": "No data found for this month"}

    # ✅ Clean numeric columns PROPERLY
    df["client_charge"] = pd.to_numeric(df["client_charge"], errors="coerce")
    df["profit"] = pd.to_numeric(df["profit"], errors="coerce")

    # 🔥 HARD FIX: remove ALL NaN completely
    df["client_charge"] = df["client_charge"].fillna(0)
    df["profit"] = df["profit"].fillna(0)

    # ✅ Dates
    df["sale_date"] = pd.to_datetime(df["sale_date"], errors="coerce")
    df["sale_date"] = df["sale_date"].dt.strftime("%Y-%m-%d")

    # ✅ Totals (SAFE)
    total_revenue = float(df["client_charge"].sum() or 0)
    total_profit = float(df["profit"].sum() or 0)

    paid = float(df.loc[df["paid_status"] == "Paid", "client_charge"].sum() or 0)
    outstanding = float(df.loc[df["paid_status"] != "Paid", "client_charge"].sum() or 0)

    # 🔥 FINAL SAFETY (NO NaN ALLOWED IN JSON)
    def safe(value):
        if pd.isna(value):
            return 0.0
        return float(value)

    data = df.to_dict(orient="records")

    # Clean each row
    for row in data:
        row["client_charge"] = safe(row.get("client_charge"))
        row["profit"] = safe(row.get("profit"))

    # ✅ Remove profit for client
    if view == "client":
        for row in data:
            row.pop("profit", None)

    return {
        "party": party,
        "month": month,
        "view": view,
        "invoices": data,
        "total_label": "Total Revenue" if view == "internal" else "Total",
        "total_value": safe(total_revenue),
        "total_profit": safe(total_profit) if view == "internal" else None,
        "paid": safe(paid),
        "outstanding": safe(outstanding)
    }

# ---------------- DASHBOARD KPIS ----------------

@app.get("/dashboard-kpis")
def dashboard_kpis():

    conn=get_conn()
    cur=conn.cursor()

    cur.execute("""
    SELECT
    COALESCE(SUM(client_charge),0) as total_sales,
    COALESCE(SUM(profit),0) as total_profit,
    COALESCE(SUM(CASE WHEN paid_status='Paid' THEN client_charge END),0) as paid_total,
    COALESCE(SUM(CASE WHEN paid_status!='Paid' THEN client_charge END),0) as outstanding_total
    FROM sales
    """)

    result=cur.fetchone()

    cur.close()
    conn.close()

    return result

@app.get("/dashboard-monthly")
def dashboard_monthly():

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
    SELECT
    TO_CHAR(sale_date,'YYYY-MM') as month,
    SUM(profit) as profit
    FROM sales
    GROUP BY month
    ORDER BY month
    """)

    rows = cur.fetchall()

    cur.close()
    conn.close()

    return rows


@app.get("/dashboard-by-party")
def dashboard_by_party():

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
    SELECT party, SUM(profit) as profit
    FROM sales
    GROUP BY party
    ORDER BY profit DESC
    """)

    rows = cur.fetchall()

    cur.close()
    conn.close()

    return rows

@app.get("/dashboard-supplier-performance")
def dashboard_supplier():

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
    SELECT supplier, SUM(profit) as profit
    FROM sales
    GROUP BY supplier
    ORDER BY profit DESC
    """)

    rows = cur.fetchall()

    cur.close()
    conn.close()

    return rows

@app.get("/dashboard-top-clients")
def dashboard_clients():

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
    SELECT party, SUM(client_charge) as revenue
    FROM sales
    GROUP BY party
    ORDER BY revenue DESC
    LIMIT 10
    """)

    rows = cur.fetchall()

    cur.close()
    conn.close()

    return rows

@app.get("/accounts-receivable")
def accounts_receivable():

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
    SELECT
    COALESCE(SUM(CASE WHEN paid_status='Paid' THEN client_charge END),0) as paid_total,
    COALESCE(SUM(CASE WHEN paid_status!='Paid' THEN client_charge END),0) as outstanding_total
    FROM sales
    """)

    result = cur.fetchone()

    cur.close()
    conn.close()

    return result


# ---------------- Export Excel ----------------
@app.get("/export-excel")
def export_excel():

    conn = get_conn()

    df = pd.read_sql("SELECT * FROM sales ORDER BY id DESC", conn)

    conn.close()

    file_path = "sales_export.xlsx"

    df.to_excel(file_path, index=False)

    return FileResponse(
        file_path,
        filename="mok_transport_sales.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ---------------- Send Monthly Report ----------------

@app.post("/send-monthly-report")
def send_monthly_report(month: str, recipient: str):

    conn = get_conn()

    df = pd.read_sql(
        f"""
        SELECT *
        FROM sales
        WHERE TO_CHAR(sale_date,'YYYY-MM')='{month}'
        """,
        conn
    )

    conn.close()

    if df.empty:
        return {"error":"No data for this month"}

    file_path = f"report_{month}.xlsx"

    df.to_excel(file_path,index=False)

    EMAIL = os.getenv("EMAIL_USER")
    PASSWORD = os.getenv("EMAIL_PASS")

    msg = EmailMessage()
    msg["Subject"] = f"Mok Transports Sales Report {month}"
    msg["From"] = EMAIL
    msg["To"] = recipient

    msg.set_content("Attached is the monthly sales report.")

    with open(file_path,"rb") as f:
        msg.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=file_path
        )

    with smtplib.SMTP_SSL("smtp.gmail.com",465) as smtp:
        smtp.login(EMAIL,PASSWORD)
        smtp.send_message(msg)

    return {"message":"Report sent"}



#---------------Revenue Forecast----------------------------------
@app.get("/revenue-forecast")
def revenue_forecast(request: Request):

    if not require_role(request, ["owner"]):
        return {"error": "owner only"}

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT sale_date, client_charge
        FROM sales
        WHERE sale_date IS NOT NULL
    """)

    rows = cur.fetchall()

    cur.close()
    conn.close()

    if not rows:
        return {"forecast": {}}

    df = pd.DataFrame(rows)

    df["sale_date"] = pd.to_datetime(df["sale_date"])
    df["client_charge"] = pd.to_numeric(df["client_charge"], errors="coerce").fillna(0)

    monthly = (
        df.groupby(df["sale_date"].dt.to_period("M"))["client_charge"]
        .sum()
        .reset_index()
    )

    monthly["month_index"] = range(len(monthly))

    x = monthly["month_index"]
    y = monthly["client_charge"]

    slope, intercept = np.polyfit(x, y, 1)

    next_month = len(monthly)

    forecast_value = slope * next_month + intercept

    return {
        "forecast_month": str((pd.Period(monthly.iloc[-1]["sale_date"]) + 1)),
        "forecast_revenue": float(forecast_value)
    }

# ---------------- OWNER DASHBOARD ----------------

@app.get("/owner-dashboard",response_class=HTMLResponse)
def owner_dashboard(request:Request):

    if not require_role(request,["owner"]):
        return RedirectResponse("/")

    conn=get_conn()
    cur=conn.cursor()

    cur.execute("""
    SELECT party,SUM(client_charge) revenue
    FROM sales
    GROUP BY party
    ORDER BY revenue DESC
    LIMIT 1
    """)

    top_client=cur.fetchone()

    cur.close()
    conn.close()

    return templates.TemplateResponse(
    request=request,
    name="owner_dashboard.html",
    context={"top_client": top_client}
     )


#------------Profit Margin Trend-------------
@app.get("/profit-margin-trend")
def profit_margin_trend(request: Request):

    if not require_role(request, ["owner"]):
        return {"error": "owner only"}

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
    SELECT sale_date, client_charge, supplier_cost
    FROM sales
    WHERE sale_date IS NOT NULL
    """)

    rows = cur.fetchall()

    cur.close()
    conn.close()

    if not rows:
        return {}

    df = pd.DataFrame(rows)

    df["sale_date"] = pd.to_datetime(df["sale_date"], errors="coerce")
    df["client_charge"] = pd.to_numeric(df["client_charge"], errors="coerce").fillna(0)
    df["supplier_cost"] = pd.to_numeric(df["supplier_cost"], errors="coerce").fillna(0)

    df["margin"] = np.where(
        df["client_charge"] > 0,
        ((df["client_charge"] - df["supplier_cost"]) / df["client_charge"]) * 100,
        0
    )

    monthly = (
        df.groupby(df["sale_date"].dt.to_period("M"))["margin"]
        .mean()
        .round(2)
    )

    result = {str(k): float(v) for k, v in monthly.items()}

    return result

#----------------Profit Alerts------------------------------
@app.get("/profit-alert")
def profit_alert(request: Request):

    if not require_role(request, ["owner"]):
        return {"error": "owner only"}

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT client_charge, supplier_cost
        FROM sales
        ORDER BY sale_date DESC
        LIMIT 20
    """)

    rows = cur.fetchall()

    cur.close()
    conn.close()

    df = pd.DataFrame(rows)

    df["client_charge"] = pd.to_numeric(df["client_charge"], errors="coerce")
    df["supplier_cost"] = pd.to_numeric(df["supplier_cost"], errors="coerce")

    df["margin"] = (df["client_charge"] - df["supplier_cost"]) / df["client_charge"] * 100

    avg_margin = df["margin"].mean()

    if avg_margin < 15:
        return {
            "alert": True,
            "message": "⚠ Profit margins are dropping below safe levels."
        }

    return {
        "alert": False,
        "message": "Profit margins healthy"
    }



@app.get("/ai-business-insights")
def ai_business_insights(request: Request):

    if not require_role(request, ["owner"]):
        return {"error":"owner only"}

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
    SELECT party, sale_date, client_charge, supplier_cost, profit
    FROM sales
    """)

    rows = cur.fetchall()

    cur.close()
    conn.close()

    if not rows:
        return {"insights":["No sales data available yet"]}

    df = pd.DataFrame(rows)

    df["sale_date"] = pd.to_datetime(df["sale_date"], errors="coerce")
    df["client_charge"] = pd.to_numeric(df["client_charge"], errors="coerce").fillna(0)
    df["profit"] = pd.to_numeric(df["profit"], errors="coerce").fillna(0)
    df["supplier_cost"] = pd.to_numeric(df["supplier_cost"], errors="coerce").fillna(0)

    insights = []

    total_revenue = df["client_charge"].sum()
    total_profit = df["profit"].sum()

    margin = (total_profit / total_revenue) * 100 if total_revenue else 0

    insights.append(f"Total revenue generated is R {total_revenue:,.2f}")
    insights.append(f"Overall profit margin is {margin:.2f}%")

    if margin < 20:
        insights.append("Profit margins are below 20%. Review supplier costs.")

    monthly = df.groupby(df["sale_date"].dt.to_period("M"))["client_charge"].sum()

    if len(monthly) > 1:

        prev = monthly.iloc[-2]
        last = monthly.iloc[-1]

        if prev > 0:
            growth = ((last - prev) / prev) * 100

            if growth > 10:
                insights.append(f"Revenue increased {growth:.1f}% last month.")

            elif growth < -10:
                insights.append(f"Revenue dropped {abs(growth):.1f}% last month.")

    top_client = df.groupby("party")["client_charge"].sum().idxmax()

    insights.append(f"Top revenue client is {top_client}")

    return {"insights": insights}


@app.get("/cashflow-forecast")
def cashflow_forecast(request: Request):

    if not require_role(request, ["owner"]):
        return {"error":"owner only"}

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
    SELECT sale_date, client_charge
    FROM sales
    WHERE sale_date IS NOT NULL
    """)

    rows = cur.fetchall()

    cur.close()
    conn.close()

    df = pd.DataFrame(rows)

    df["sale_date"] = pd.to_datetime(df["sale_date"])
    df["client_charge"] = pd.to_numeric(df["client_charge"], errors="coerce")

    monthly = (
        df.groupby(df["sale_date"].dt.to_period("M"))["client_charge"]
        .sum()
        .reset_index()
    )

    monthly["index"] = range(len(monthly))

    x = monthly["index"]
    y = monthly["client_charge"]

    slope, intercept = np.polyfit(x, y, 1)

    forecasts = {}

    for i in range(1,4):

        month = str(monthly.iloc[-1]["sale_date"] + i)

        forecasts[month] = float(slope*(len(monthly)+i)+intercept)

    return forecasts

@app.get("/logistics-profit-map")
def logistics_profit_map(request: Request):

    if not require_role(request, ["owner"]):
        return {"error":"owner only"}

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
    SELECT supplier, party, SUM(profit) as profit
    FROM sales
    GROUP BY supplier, party
    """)

    rows = cur.fetchall()

    cur.close()
    conn.close()

    df = pd.DataFrame(rows)

    result = {}

    for _, row in df.iterrows():

        supplier = row["supplier"]

        if supplier not in result:
            result[supplier] = {}

        result[supplier][row["party"]] = float(row["profit"])

    return result


@app.post("/add-party")
def add_party(name:str):

 if name and name not in PARTIES:
  PARTIES.append(name)

 return {"status":"added","party":name}

@app.get("/late-payment-alerts")
def late_payment_alerts(request: Request):

    if not require_role(request, ["owner","accounts"]):
        return {"error":"permission denied"}

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
    SELECT party, invoice_no, sale_date, client_charge
    FROM sales
    WHERE paid_status != 'Paid'
    AND sale_date IS NOT NULL
    """)

    rows = cur.fetchall()

    cur.close()
    conn.close()

    if not rows:
        return {"alerts":[]}

    df = pd.DataFrame(rows)

    df["sale_date"] = pd.to_datetime(df["sale_date"], errors="coerce")
    df["client_charge"] = pd.to_numeric(df["client_charge"], errors="coerce").fillna(0)

    df["days"] = (pd.Timestamp.today() - df["sale_date"]).dt.days

    late = df[df["days"] > 30]

    alerts = []

    for _, row in late.iterrows():

        alerts.append({
            "party": row["party"],
            "invoice": row["invoice_no"],
            "days": int(row["days"]),
            "amount": float(row["client_charge"])
        })

    return {"alerts": alerts}



@app.get("/client-growth-opportunities")
def client_growth_opportunities(request: Request):

    if not require_role(request, ["owner"]):
        return {"error":"owner only"}

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
    SELECT party, sale_date, client_charge
    FROM sales
    WHERE sale_date IS NOT NULL
    """)

    rows = cur.fetchall()

    cur.close()
    conn.close()

    if not rows:
        return {"clients":[]}

    df = pd.DataFrame(rows)

    df["sale_date"] = pd.to_datetime(df["sale_date"])
    df["client_charge"] = pd.to_numeric(df["client_charge"], errors="coerce").fillna(0)

    df["month"] = df["sale_date"].dt.to_period("M")

    grouped = df.groupby(["party","month"])["client_charge"].sum().reset_index()

    growth_clients = []

    for client in grouped["party"].unique():

        client_data = grouped[grouped["party"] == client]

        if len(client_data) < 2:
            continue

        last = client_data.iloc[-1]["client_charge"]
        prev = client_data.iloc[-2]["client_charge"]

        if prev > 0:

            growth = ((last - prev) / prev) * 100

            if growth > 20:

                growth_clients.append({
                    "client": client,
                    "growth": round(growth,1)
                })

    return {"clients": growth_clients}


# ---------------- OWNER ANALYTICS ----------------
@app.get("/owner-analytics")
def owner_analytics(request: Request):

    if not require_role(request, ["owner"]):
        return {"error": "owner only"}

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("SELECT * FROM sales")
    rows = cur.fetchall()

    cur.close()
    conn.close()

    if not rows:
        return {
            "revenue": 0,
            "profit": 0,
            "monthly_profit": {},
            "client_profit": {},
            "supplier_profit": {}
        }

    df = pd.DataFrame(rows)

    df["sale_date"] = pd.to_datetime(df["sale_date"], errors="coerce")
    df["client_charge"] = pd.to_numeric(df["client_charge"], errors="coerce").fillna(0)
    df["profit"] = pd.to_numeric(df["profit"], errors="coerce").fillna(0)

    revenue = float(df["client_charge"].sum())
    profit = float(df["profit"].sum())

    df_clean = df.dropna(subset=["sale_date"])

    monthly_profit = (
        df_clean.groupby(df_clean["sale_date"].dt.to_period("M"))["profit"]
        .sum()
        .astype(float)
        .to_dict()
    )

    monthly_profit = {str(k): v for k, v in monthly_profit.items()}

    client_profit = (
        df.groupby("party")["profit"]
        .sum()
        .astype(float)
        .to_dict()
    )

    supplier_profit = (
        df.groupby("supplier")["profit"]
        .sum()
        .astype(float)
        .to_dict()
    )

    return {
        "revenue": revenue,
        "profit": profit,
        "monthly_profit": monthly_profit,
        "client_profit": client_profit,
        "supplier_profit": supplier_profit
    }


#-------------High Profit Clients----------------

@app.get("/high-profit-clients")
def high_profit_clients(request: Request):

    if not require_role(request, ["owner"]):
        return {"error":"owner only"}

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
    SELECT party, SUM(profit) as total_profit
    FROM sales
    GROUP BY party
    ORDER BY total_profit DESC
    LIMIT 5
    """)

    rows = cur.fetchall()

    cur.close()
    conn.close()

    return {"clients": rows}

#-------------Client Concentration Risk------
@app.get("/client-concentration-risk")
def client_concentration_risk(request: Request):

    if not require_role(request, ["owner"]):
        return {"error":"owner only"}

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
    SELECT party, SUM(client_charge) as revenue
    FROM sales
    GROUP BY party
    """)

    rows = cur.fetchall()

    cur.close()
    conn.close()

    if not rows:
        return {"risk":None}

    df = pd.DataFrame(rows)

    total = df["revenue"].sum()

    df["share"] = (df["revenue"] / total) * 100

    top_client = df.sort_values("share", ascending=False).iloc[0]

    return {
        "client": top_client["party"],
        "share": round(top_client["share"],1)
    }

#------------Client Target Engine----
@app.get("/ai-client-targeting")
def ai_client_targeting(request: Request):

    if not require_role(request, ["owner"]):
        return {"error":"owner only"}

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
    SELECT party, sale_date, client_charge
    FROM sales
    WHERE sale_date IS NOT NULL
    """)

    rows = cur.fetchall()

    cur.close()
    conn.close()

    if not rows:
        return {"targets":[]}

    df = pd.DataFrame(rows)

    df["sale_date"] = pd.to_datetime(df["sale_date"], errors="coerce")
    df["client_charge"] = pd.to_numeric(df["client_charge"], errors="coerce").fillna(0)

    df["month"] = df["sale_date"].dt.to_period("M")

    grouped = df.groupby(["party","month"])["client_charge"].sum().reset_index()

    insights = []

    for client in grouped["party"].unique():

        client_data = grouped[grouped["party"] == client]

        if len(client_data) < 2:
            continue

        prev = client_data.iloc[-2]["client_charge"]
        last = client_data.iloc[-1]["client_charge"]

        if prev == 0:
            continue

        change = ((last - prev) / prev) * 100

        if change > 25:

            insights.append({
                "client": client,
                "type": "growth",
                "message": f"{client} shipments growing {round(change,1)}%. Target for bigger contract."
            })

        elif change < -25:

            insights.append({
                "client": client,
                "type": "decline",
                "message": f"{client} shipments declining {round(abs(change),1)}%. Risk of losing client."
            })

    return {"targets": insights}


#------------Client Statements----

@app.get("/export-client-statement")
def export_client_statement(party: str, month: str, view: str = "internal"):

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT invoice_no, sale_date, client_charge, profit, paid_status
        FROM sales
        WHERE LOWER(party) = LOWER(%s)
        AND sale_date IS NOT NULL
        AND TO_CHAR(sale_date,'YYYY-MM') = %s
        ORDER BY sale_date
    """, (party, month))

    rows = cur.fetchall()

    cur.close()
    conn.close()

    if not rows:
        return {"error": "No data"}

    df = pd.DataFrame(rows)

    df["client_charge"] = pd.to_numeric(df["client_charge"], errors="coerce").fillna(0)
    df["profit"] = pd.to_numeric(df["profit"], errors="coerce").fillna(0)

    total_revenue = df["client_charge"].sum()
    total_profit = df["profit"].sum()
    paid = df[df["paid_status"] == "Paid"]["client_charge"].sum()
    outstanding = df[df["paid_status"] != "Paid"]["client_charge"].sum()

    wb = Workbook()
    ws = wb.active
    ws.title = "Statement"

    # Header
    ws["A1"] = "Mok Transports"
    ws["A2"] = "Client Statement"
    ws["A3"] = f"Client: {party}"
    ws["A4"] = f"Month: {month}"

    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"].font = Font(size=14, bold=True)

    headers = ["Invoice", "Date", "Amount", "Status"]

    if view == "internal":
        headers.insert(3, "Profit")

    ws.append([])
    ws.append(headers)

    for _, row in df.iterrows():

        row_data = [
            row["invoice_no"],
            str(row["sale_date"]),
            float(row["client_charge"]),
            row["paid_status"]
        ]

        if view == "internal":
            row_data.insert(3, float(row["profit"]))

        ws.append(row_data)

    ws.append([])
    ws.append(["Total Revenue", total_revenue])
    if view == "internal":
        ws.append(["Total Profit", total_profit])
    ws.append(["Paid", paid])
    ws.append(["Outstanding", outstanding])

    file_path = f"{party}_{month}.xlsx"
    wb.save(file_path)

    return FileResponse(file_path, filename=file_path)

#------------Email Statements----

@app.post("/email-client-statement")
def email_client_statement(party: str, month: str, email: str = ""):

    conn = get_conn()

    query = """
        SELECT invoice_no, sale_date, client_charge, profit, paid_status
        FROM sales
        WHERE LOWER(party) = LOWER(%s)
        AND sale_date IS NOT NULL
        AND TO_CHAR(sale_date,'YYYY-MM') = %s
        ORDER BY sale_date
    """

    df = pd.read_sql(query, conn, params=(party, month))
    conn.close()

    if df.empty:
        return {"error": "No data"}

    file_path = f"{party}_{month}.xlsx"
    df.to_excel(file_path, index=False)

    EMAIL = os.getenv("EMAIL_USER")
    PASSWORD = os.getenv("EMAIL_PASS")

    msg = EmailMessage()
    msg["Subject"] = f"{party} Statement {month}"
    msg["From"] = EMAIL

    recipients = ["ryan@moktransports.com"]  

    if email:
        recipients.append(email)

    msg["To"] = ", ".join(recipients)

    msg.set_content(f"Attached is the statement for {party} - {month}")

    with open(file_path, "rb") as f:
        msg.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=file_path
        )

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(EMAIL, PASSWORD)
        smtp.send_message(msg)

    return {"message": "Statement emailed successfully"}



if __name__=="__main__":

    port=int(os.environ.get("PORT",8080))

    uvicorn.run(
        app,
        host="0.0.0.0",
        port=port
    )






 
 